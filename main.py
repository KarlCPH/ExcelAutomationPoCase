import pandas as pd
import numpy as np
from classes import Product
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from mapping import PRODUCT_ID, PRODUCT_MOLECULE, PRODUCT_PACK_SIZE, PRODUCT_PACK_QUANTITY, PRODUCT_TREATMENT

wb = load_workbook(filename="POcase.xlsx")
wb2 = load_workbook(filename="HELLO.xlsx")
sheet = wb.active
sheet2 = wb2.active

ws = wb["DPPIV & SGLT2"]

sheet2["A5"] = "IM SO COOL"

wb2.save("HELLO.xlsx")

products = []

for row in sheet.iter_rows(min_row=8, max_row=3030, max_col=10, values_only=True):
    product = Product(id=row[PRODUCT_ID],
                      Molecule=row[PRODUCT_MOLECULE],
                      Pack_Size=row[PRODUCT_PACK_SIZE],
                      Pack_Quantity=row[PRODUCT_PACK_QUANTITY],
                      Product_Treatment=row[PRODUCT_TREATMENT])
    products.append(product)

print(products[1000])

df = pd.DataFrame(ws.values)

df = df.drop(labels=6, axis=0)

df = df.drop(labels=5, axis=0)

df[5] = pd.to_numeric(df[5])

df[7] = pd.to_numeric(df[7])

df[12] = df[7] * df[5] / 1

comparison_column = np.where(df[12] == df[8], True, False)

print(df[12])










